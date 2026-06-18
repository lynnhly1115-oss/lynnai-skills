#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import subprocess
from collections import OrderedDict, defaultdict
from pathlib import Path

try:
    from docx import Document
except ImportError:  # pragma: no cover - dependency depends on the active runtime
    Document = None

try:
    import pdfplumber
except ImportError:  # pragma: no cover - dependency depends on the active runtime
    pdfplumber = None

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DAY_WORDS = "日|夜|清晨|上午|下午|傍晚|黄昏|深夜|凌晨|白天|晚上"
IO_WORDS = "内外|内|外"
INSERT_KEYWORDS = ("一组镜头", "手机屏幕", "蒙太奇", "转场", "插入", "空镜", "闪回", "闪现")
GROUP_HINTS = ("群众", "众人", "村民", "员工", "游客", "工作人员", "代表", "保安", "保镖", "学生", "干部", "客人", "路人", "若干")


CN_DIGITS = {
    "零": 0,
    "一": 1,
    "二": 2,
    "两": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
}


PROP_RULES = [
    {
        "name": "手机/屏幕画面",
        "keywords": ("手机", "电话", "屏幕", "视频", "通话", "微信", "短信", "消息"),
        "category": "关键剧情/特写图文",
        "purpose": "通话、信息展示或屏幕特写",
    },
    {
        "name": "直播设备",
        "keywords": ("直播", "支架", "补光灯", "麦克风", "收音麦"),
        "category": "特殊道具/电子设备",
        "purpose": "直播、拍摄或传播画面",
    },
    {
        "name": "文件/合同/台账",
        "keywords": ("文件", "合同", "协议", "台账", "登记本", "账本", "名单", "报表", "证明", "证书", "通知书"),
        "category": "关键剧情/图文道具",
        "purpose": "信息交代、签署、举证或特写",
    },
    {
        "name": "现金/红包/银行卡",
        "keywords": ("现金", "钱", "红包", "银行卡", "转账", "分红", "补偿款"),
        "category": "关键剧情/财物",
        "purpose": "财务分配、补偿或交易",
    },
    {
        "name": "牌匾/招牌/门牌",
        "keywords": ("牌匾", "招牌", "门牌", "标识", "标牌", "牌子", "横幅", "锦旗"),
        "category": "图文道具/置景标识",
        "purpose": "空间身份、剧情事件或视觉记忆点",
    },
    {
        "name": "医药用品",
        "keywords": ("药", "药箱", "药瓶", "针", "输液", "病历", "处方", "诊断", "听诊器", "血压计", "纱布", "绷带"),
        "category": "专业道具",
        "purpose": "诊疗、救治或身体状态表现",
    },
    {
        "name": "会议用品",
        "keywords": ("会议", "话筒", "桌牌", "茶杯", "水杯", "投影", "白板"),
        "category": "陈设/普通道具",
        "purpose": "会议、讨论、公共空间陈设",
    },
    {
        "name": "餐饮食物",
        "keywords": ("吃饭", "饭菜", "包子", "酒", "喝茶", "茶杯", "水杯", "碗", "筷", "小菜", "土酒", "饮料"),
        "category": "消耗道具",
        "purpose": "生活烟火气、饭局或饮食动作",
    },
    {
        "name": "行李/包袋",
        "keywords": ("行李", "箱子", "包", "背包", "提包", "袋子"),
        "category": "普通道具",
        "purpose": "人物出入、搬离或身份状态",
    },
    {
        "name": "施工修缮工具",
        "keywords": ("施工", "修缮", "锤", "钉", "梯子", "工具", "木板", "油漆", "刷子", "围挡"),
        "category": "特殊道具/置景施工",
        "purpose": "修建、翻新、破损修复或工程现场",
    },
    {
        "name": "灯具/油灯",
        "keywords": ("灯泡", "灯", "油灯", "手电", "蜡烛"),
        "category": "陈设/实用光源",
        "purpose": "生活环境、夜景氛围或特写",
    },
    {
        "name": "照片/相框",
        "keywords": ("照片", "相框", "合影", "画像"),
        "category": "关键剧情/图文道具",
        "purpose": "关系说明、回忆或空间生活痕迹",
    },
    {
        "name": "农具/乡村用具",
        "keywords": ("锄头", "扁担", "竹篮", "箩筐", "扫帚", "水桶", "鸡鸭"),
        "category": "陈设/普通道具",
        "purpose": "乡村生活质感和空间可信度",
    },
    {
        "name": "车辆/车钥匙",
        "keywords": ("车", "汽车", "面包车", "车钥匙", "司机"),
        "category": "交通道具",
        "purpose": "出行、到达或身份区分",
    },
    {
        "name": "活物/动物",
        "keywords": ("活物", "动物", "鸡鸭", "鸡群", "鸭群", "猫", "狗", "鹅", "猪", "牛羊", "羊群"),
        "category": "特殊道具/活物",
        "purpose": "活物调度、乡村生活或剧情需要",
    },
    {
        "name": "危险/破损道具",
        "keywords": ("刀", "棍", "砸", "摔碎", "打碎", "破损", "血", "着火", "烫", "水泼"),
        "category": "危险动作/损坏道具",
        "purpose": "冲突、破坏、受伤或动作安全",
    },
]


LOCATION_DRESSING = [
    {
        "keywords": ("村委会", "会议室", "办公室"),
        "wall": "基层办公墙面、公示栏、制度牌、门牌",
        "furniture": "会议桌、椅子、桌牌、茶水区",
        "soft": "文件夹、笔、本、热水杯、旧办公痕迹",
        "graphic": "村委会标识、会议资料、公告文字",
    },
    {
        "keywords": ("老宅", "祖宅", "家", "屋", "房间", "院"),
        "wall": "旧墙面、门窗、地面磨损、院落入口",
        "furniture": "木桌椅、床、柜、灶台或生活家具",
        "soft": "旧布料、生活杂物、照片、碗筷、水壶",
        "graphic": "门牌、旧照片、手写纸条",
    },
    {
        "keywords": ("医院", "诊所", "医馆", "病房"),
        "wall": "医疗白墙、科室牌、帘布、洁净地面",
        "furniture": "病床、诊疗桌、药柜、输液架",
        "soft": "病历、药瓶、纱布、诊疗用品",
        "graphic": "病历、处方、检查单需专业核对",
    },
    {
        "keywords": ("桥", "福星桥"),
        "wall": "桥体、栏杆、桥面、桥头地面",
        "furniture": "施工围挡、工具堆、临时指示牌",
        "soft": "修缮痕迹、灰尘、旧化和翻新对比",
        "graphic": "桥名标识、施工告示、文旅标识",
    },
    {
        "keywords": ("民宿", "客栈", "酒店"),
        "wall": "前台背景墙、房门、走廊、门牌",
        "furniture": "前台、沙发、行李架、客房家具",
        "soft": "床品、绿植、摆件、客用品",
        "graphic": "民宿招牌、房卡、入住登记",
    },
    {
        "keywords": ("餐馆", "饭店", "厨房", "包厢"),
        "wall": "店招、菜单墙、后厨墙面、油烟痕迹",
        "furniture": "餐桌椅、吧台、餐具柜",
        "soft": "碗筷、酒水、菜品、纸巾、桌布",
        "graphic": "菜单、价目表、店铺招牌",
    },
    {
        "keywords": ("路口", "村口", "街", "广场", "田", "山"),
        "wall": "村路、路牌、墙面标语、自然环境",
        "furniture": "长椅、摊位、临时围观区域",
        "soft": "农具、竹篮、生活杂物、车辆",
        "graphic": "路牌、村名牌、宣传标语",
    },
    {
        "keywords": ("直播", "手机屏幕", "屏幕"),
        "wall": "以画面内容为主，需准备可拍屏幕或后期图",
        "furniture": "直播支架、补光灯、桌面",
        "soft": "产品、包装、发货物料",
        "graphic": "直播界面、弹幕、账号名、商品图文需审核",
    },
]


def cn_to_int(value: str) -> int | None:
    value = value.strip()
    if value.isdigit():
        return int(value)
    if value in CN_DIGITS:
        return CN_DIGITS[value]
    if value == "十":
        return 10
    if "十" in value:
        left, _, right = value.partition("十")
        tens = 1 if not left else CN_DIGITS.get(left)
        ones = 0 if not right else CN_DIGITS.get(right)
        if tens is None or ones is None:
            return None
        return tens * 10 + ones
    return None


def normalize_day(day: str) -> str:
    if day in {"清晨", "上午", "下午", "傍晚", "黄昏", "白天"}:
        return "日"
    if day in {"深夜", "凌晨", "晚上"}:
        return "夜"
    if day in {"日/夜", "夜/日"}:
        return "日/夜"
    return day


def compact_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text.replace("\u3000", " ")).strip()


def normalize_for_heading(text: str) -> str:
    return re.sub(r"\s+", "", text.replace("\u3000", " "))


def clean_source_line(text: str) -> str:
    text = compact_spaces(text)
    text = re.sub(r"^\s*[△▲]\s*", "", text)
    return text.strip()


def ordered_unique(items: list[str]) -> list[str]:
    seen = OrderedDict()
    for item in items:
        item = item.strip(" 、，,;；")
        if item:
            seen[item] = True
    return list(seen.keys())


def split_people(line: str) -> list[str]:
    line = re.sub(r"^人物\s*[:：;；]?", "", line.strip())
    line = re.sub(r"[（(][^）)]*[）)]", "", line)
    parts = re.split(r"[、，,;；/\s]+", line)
    result = []
    for part in parts:
        part = part.strip(" :：")
        if part and part not in {"人物", "无"}:
            result.append(part)
    return ordered_unique(result)


def detect_episode(text: str) -> int | str | None:
    stripped = normalize_for_heading(text)
    if stripped.startswith("番外篇") or stripped == "番外":
        return "番外"
    match = re.match(r"^第([一二两三四五六七八九十\d]+)[集幕章]", stripped)
    if match:
        return cn_to_int(match.group(1))
    return None


def looks_like_heading(text: str) -> bool:
    stripped = normalize_for_heading(text)
    day = f"(?:{DAY_WORDS}|日/夜|夜/日)"
    io = f"(?:{IO_WORDS})"
    patterns = [
        rf"^\d+[-—]\d+(?:{day})?{io}.+$",
        rf"^\d+[.、].+?{day}{io}$",
        rf"^{day}{io}.+$",
        rf"^第?\d+场[:：]?.+?{day}{io}$",
        r"^\d+[.、].*(一组镜头|手机屏幕|蒙太奇|转场|插入|空镜|闪回|闪现).*$",
    ]
    return any(re.match(pattern, stripped) for pattern in patterns)


def parse_heading(
    text: str,
    episode: int | str | None,
    scene_count: dict[int | str, int],
    last_day: str,
    last_io: str,
    sequence: int,
) -> dict | None:
    stripped = normalize_for_heading(text)
    day_re = f"({DAY_WORDS}|日/夜|夜/日)"
    io_re = f"({IO_WORDS})"

    match = re.match(rf"^(\d+)[-—](\d+)({day_re})?{io_re}(.+)$", stripped)
    if match:
        ep = int(match.group(1))
        no = int(match.group(2))
        day = normalize_day(match.group(3) or last_day or "日")
        scene_count[ep] = max(scene_count.get(ep, 0), no)
        return {
            "scene": f"{ep}-{no}",
            "source_num": str(no),
            "day": day,
            "inside_outside": match.group(4),
            "location": match.group(5).strip(),
            "warning": "" if match.group(3) else "原文未标日夜，沿用前场",
        }

    match = re.match(rf"^(\d+)[.、](.+?){day_re}{io_re}$", stripped)
    if match:
        source_num = int(match.group(1))
        location = match.group(2).strip()
        day = normalize_day(match.group(3))
        inside_outside = match.group(4)
        if episode is not None:
            scene_count[episode] = scene_count.get(episode, 0) + 1
            scene_id = f"{episode}-{scene_count[episode]}" if episode != "番外" else f"番外-{scene_count[episode]}"
            warning = "" if scene_count[episode] == source_num or episode == "番外" else f"原文场号{source_num}"
        else:
            scene_id = str(source_num)
            warning = ""
        return {
            "scene": scene_id,
            "source_num": str(source_num),
            "day": day,
            "inside_outside": inside_outside,
            "location": location,
            "warning": warning,
        }

    match = re.match(rf"^第?(\d+)场[:：]?(.+?){day_re}{io_re}$", stripped)
    if match:
        source_num = int(match.group(1))
        location = match.group(2).strip()
        day = normalize_day(match.group(3))
        inside_outside = match.group(4)
        scene_id = f"{episode}-{source_num}" if episode not in {None, '番外'} else str(source_num)
        return {
            "scene": scene_id,
            "source_num": str(source_num),
            "day": day,
            "inside_outside": inside_outside,
            "location": location,
            "warning": "",
        }

    match = re.match(rf"^{day_re}{io_re}(.+)$", stripped)
    if match and episode is not None:
        scene_count[episode] = scene_count.get(episode, 0) + 1
        scene_id = f"{episode}-{scene_count[episode]}" if episode != "番外" else f"番外-{scene_count[episode]}"
        return {
            "scene": scene_id,
            "source_num": str(scene_count[episode]),
            "day": normalize_day(match.group(1)),
            "inside_outside": match.group(2),
            "location": match.group(3).strip(),
            "warning": "原文无显式场号",
        }

    match = re.match(r"^(\d+)[.、](.+)$", stripped)
    if match and any(keyword in match.group(2) for keyword in INSERT_KEYWORDS):
        source_num = int(match.group(1))
        location = match.group(2).strip()
        if episode is not None:
            scene_count[episode] = scene_count.get(episode, 0) + 1
            scene_id = f"{episode}-{scene_count[episode]}" if episode != "番外" else f"番外-{scene_count[episode]}"
            warning = "原文未标日夜内外"
            if scene_count[episode] != source_num and episode != "番外":
                warning += f"；原文场号{source_num}"
        else:
            scene_id = str(sequence)
            warning = "原文未标日夜内外"
        return {
            "scene": scene_id,
            "source_num": str(source_num),
            "day": last_day or "日",
            "inside_outside": "内外" if any(k in location for k in ("一组镜头", "蒙太奇", "转场")) else (last_io or "内"),
            "location": location,
            "warning": warning,
        }

    return None


def pdf_lines(path: Path) -> list[dict]:
    if pdfplumber is None:
        raise RuntimeError("pdfplumber is required to parse PDF scripts")

    lines: list[dict] = []
    with pdfplumber.open(path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            words = page.extract_words(
                extra_attrs=["fontname", "size", "non_stroking_color"],
                keep_blank_chars=False,
            )
            filtered = [
                word
                for word in words
                if word.get("size", 0) <= 20 and word.get("fontname") != "STSong"
            ]
            groups: list[dict] = []
            for word in filtered:
                for group in groups:
                    if abs(group["top"] - word["top"]) < 3:
                        group["words"].append(word)
                        group["top"] = (group["top"] + word["top"]) / 2
                        break
                else:
                    groups.append({"top": word["top"], "words": [word]})
            for group in sorted(groups, key=lambda item: item["top"]):
                text = "".join(
                    word["text"] for word in sorted(group["words"], key=lambda item: item["x0"])
                ).strip()
                if text:
                    lines.append({"text": text, "page": page_number})

    if not lines:
        raise RuntimeError("No selectable text found in PDF. This looks like a scanned/image-only PDF; OCR or a Word/text export is required.")
    return lines


def docx_lines(path: Path) -> list[dict]:
    if Document is None:
        raise RuntimeError("python-docx is required to parse DOCX scripts")
    return [
        {"text": paragraph.text.strip(), "page": ""}
        for paragraph in Document(path).paragraphs
        if paragraph.text.strip()
    ]


def doc_lines(path: Path) -> list[dict]:
    try:
        proc = subprocess.run(
            ["textutil", "-convert", "txt", "-stdout", str(path)],
            check=True,
            capture_output=True,
            text=True,
        )
    except (FileNotFoundError, subprocess.CalledProcessError) as exc:
        raise RuntimeError(f"Could not extract .doc text with textutil: {exc}") from exc
    return [
        {"text": line.strip(), "page": ""}
        for line in proc.stdout.splitlines()
        if line.strip()
    ]


def load_lines(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return pdf_lines(path)
    if suffix in {".docx", ".docm"}:
        return docx_lines(path)
    if suffix == ".doc":
        return doc_lines(path)
    raise ValueError(f"Unsupported input format: {path.suffix}")


def collect_people_line(lines: list[dict], start: int, end: int) -> str:
    for idx in range(start + 1, min(end, start + 10)):
        text = clean_source_line(lines[idx]["text"])
        if text.startswith("人物"):
            people_line = text
            cursor = idx + 1
            while cursor < min(end, idx + 5):
                next_text = clean_source_line(lines[cursor]["text"])
                if looks_like_heading(next_text) or detect_episode(next_text) is not None:
                    break
                if (
                    people_line.endswith(("、", "，", ","))
                    or (
                        len(next_text) <= 18
                        and "：" not in next_text
                        and ":" not in next_text
                        and not re.search(r"[。！？；;]", next_text)
                    )
                ):
                    people_line += next_text
                    cursor += 1
                    continue
                break
            return people_line
        if looks_like_heading(text) or detect_episode(text) is not None:
            break
    return ""


def parse_script(path: Path) -> list[dict]:
    lines = load_lines(path)
    scenes: list[dict] = []
    episode: int | str | None = None
    scene_count: dict[int | str, int] = {}
    last_day = "日"
    last_io = "内"
    sequence = 1
    in_body = not any(clean_source_line(line["text"]).startswith("正文") for line in lines)

    for idx, line in enumerate(lines):
        text = clean_source_line(line["text"])
        if text.startswith("正文"):
            in_body = True
            continue
        if not in_body:
            continue
        detected_episode = detect_episode(text)
        if detected_episode is not None:
            episode = detected_episode
            scene_count.setdefault(episode, 0)
            continue
        if not looks_like_heading(text):
            continue
        parsed = parse_heading(text, episode, scene_count, last_day, last_io, sequence)
        if not parsed:
            continue
        parsed.update(
            {
                "source": path.name,
                "idx": idx,
                "heading": text,
                "source_page": line.get("page", ""),
            }
        )
        scenes.append(parsed)
        last_day = parsed["day"] if parsed["day"] != "日/夜" else last_day
        last_io = parsed["inside_outside"]
        sequence += 1

    for scene_index, scene in enumerate(scenes):
        end = scenes[scene_index + 1]["idx"] if scene_index + 1 < len(scenes) else len(lines)
        people_line = collect_people_line(lines, scene["idx"], end)
        body = []
        for line in lines[scene["idx"] + 1 : end]:
            text = clean_source_line(line["text"])
            if not text:
                continue
            if text.startswith("人物") or detect_episode(text) is not None or text.startswith("正文"):
                continue
            if looks_like_heading(text):
                continue
            body.append(text)
        scene["people_raw"] = people_line
        scene["people"] = split_people(people_line)
        scene["body"] = body
        scene["body_text"] = "\n".join(body)

    return scenes


def clean_action_text(line: str) -> str:
    line = re.sub(r"^△", "", line).strip()
    line = re.sub(r"^[A-Za-z0-9_\-]+[:：]", "", line).strip()
    line = re.sub(r"^[\u4e00-\u9fa5·]{1,8}[:：]", "", line).strip()
    line = re.sub(r"【[^】]*】", "", line)
    line = re.sub(r"[（(][^）)]{0,30}[）)]", "", line)
    line = re.sub(r"\s+", "", line)
    return line.strip(" ，。！？；;：:")


def short_text(text: str, limit: int) -> str:
    text = re.sub(r"\s+", "", text)
    if len(text) <= limit:
        return text
    return text[:limit]


def scene_summary(scene: dict) -> str:
    candidates = []
    for line in scene["body"]:
        cleaned = clean_action_text(line)
        if not cleaned:
            continue
        if len(cleaned) >= 6:
            candidates.append(cleaned)
        if line.startswith(("△", "▲")) and cleaned:
            candidates.insert(0, cleaned)
    if candidates:
        summary = short_text(candidates[0], 24)
    elif any(keyword in scene["location"] for keyword in INSERT_KEYWORDS):
        summary = short_text(f"{scene['location']}交代剧情信息", 24)
    else:
        summary = short_text(f"{scene['location']}发生剧情推进", 24)
    if len(summary) < 8 and scene["location"]:
        summary = short_text(f"{scene['location']}{summary}", 24)
    return summary


def people_and_groups(scene: dict) -> str:
    text = scene["people_raw"] + "\n" + scene["body_text"]
    names = list(scene["people"])
    for hint in GROUP_HINTS:
        if hint in text and hint not in names:
            names.append(hint)
    return "、".join(ordered_unique(names)) or "人物未列明（需确认）"


def is_group_name(name: str) -> bool:
    return any(hint in name for hint in GROUP_HINTS)


def named_people(scene: dict) -> list[str]:
    return [name for name in scene.get("people", []) if not is_group_name(name)]


def normalize_location(location: str) -> str:
    location = re.sub(r"[（(].*?[）)]", "", location)
    location = re.sub(r"(接\d+[-—]\d+|同场|一同拍摄)", "", location)
    location = location.strip(" -—/、，,;；:：")
    return location or "特殊画面"


def normalize_subspace(subspace: str, inside_outside: str = "") -> str:
    subspace = subspace.strip(" -—/、，,;；:：")
    if not subspace:
        return "内景" if inside_outside == "内" else ("外景" if inside_outside == "外" else "未细分空间")
    if subspace == "外":
        return "外景/院外" if inside_outside == "外" else "外景"
    if subspace == "内":
        return "内景"
    if any(word in subspace for word in ("客厅", "厅")):
        return "客厅"
    if "厨房" in subspace or "灶" in subspace:
        return "厨房"
    if any(word in subspace for word in ("卧室", "房间", "屋内", "屋里")):
        return "卧室/房间"
    if any(word in subspace for word in ("院子", "院内", "院外", "前院", "后院", "院")):
        return subspace
    if any(word in subspace for word in ("门口", "门外", "门内", "门前")):
        return subspace
    if any(word in subspace for word in ("会议室门口", "会议室外", "会议室内")):
        return subspace
    if "会议室" in subspace:
        return "会议室"
    if "办公室" in subspace:
        return "办公室"
    if "病房" in subspace:
        return "病房"
    if "桥上" in subspace or subspace == "上":
        return "桥上"
    if "屏幕" in subspace:
        return "屏幕画面"
    return subspace


def default_subspace(main_set: str, inside_outside: str = "") -> str:
    if inside_outside == "内":
        if main_set.endswith("家") or any(word in main_set for word in ("老宅", "祖宅")):
            return "家中内景"
        return f"{main_set}内景"
    if inside_outside == "外":
        if main_set.endswith("家") or any(word in main_set for word in ("老宅", "祖宅")):
            return "家外景"
        return f"{main_set}外景"
    return f"{main_set}未细分空间"


def finalize_subspace(main_set: str, subspace: str, inside_outside: str = "") -> str:
    if subspace in {"内景", "外景", "未细分空间"}:
        return default_subspace(main_set, inside_outside)
    return subspace


def split_location_for_scouting(location: str, inside_outside: str = "") -> tuple[str, str]:
    loc = normalize_location(location)
    if any(keyword in loc for keyword in INSERT_KEYWORDS):
        main = "特殊画面/插入镜头"
        return main, finalize_subspace(main, normalize_subspace(loc, inside_outside), inside_outside)

    office_match = re.match(r"^(.+?)办公室$", loc)
    if office_match:
        prefix = office_match.group(1)
        main = prefix if prefix in {"纪委", "村委会"} else f"{prefix}办公区"
        return main, "办公室"

    ordered_markers = (
        "村委会",
        "张家老宅",
        "王老太老宅",
        "老宅",
        "祖宅",
        "民宿",
        "客栈",
        "酒店",
        "医院",
        "诊所",
        "医馆",
        "餐馆",
        "饭店",
        "办公室",
        "宿舍",
        "福星桥",
        "桥",
    )
    for marker in ordered_markers:
        idx = loc.find(marker)
        if idx >= 0:
            main = loc[: idx + len(marker)]
            sub = loc[idx + len(marker) :]
            if marker == "村委会" and not sub:
                sub = ""
            return main, finalize_subspace(main, normalize_subspace(sub, inside_outside), inside_outside)

    home_match = re.match(r"^(.+?家)(.*)$", loc)
    if home_match:
        main = home_match.group(1)
        sub = home_match.group(2)
        return main, finalize_subspace(main, normalize_subspace(sub, inside_outside), inside_outside)

    yard_match = re.match(r"^(.+?)院(内|外)?$", loc)
    if yard_match and yard_match.group(1):
        main = f"{yard_match.group(1)}院"
        sub = f"院{yard_match.group(2)}" if yard_match.group(2) else "院子"
        return main, finalize_subspace(main, normalize_subspace(sub, inside_outside), inside_outside)

    if loc in {"院子", "院内", "院外"}:
        main = "未明确宅院（需确认）"
        return main, finalize_subspace(main, normalize_subspace(loc, inside_outside), inside_outside)

    if "路口" in loc:
        return loc, f"{loc}外景"
    if "村口" in loc:
        return loc, f"{loc}外景"
    return loc, finalize_subspace(loc, normalize_subspace("", inside_outside), inside_outside)


def match_location_rule(location: str) -> dict | None:
    for rule in LOCATION_DRESSING:
        if any(keyword in location for keyword in rule["keywords"]):
            return rule
    return None


def extract_live_animals(scene: dict) -> list[str]:
    text = "\n".join(scene.get("body", []))
    text = text.replace("阿猫阿狗", "")
    animals = []
    animal_patterns = [
        ("鸡", r"(?:几只鸡|一只鸡|两只鸡|这批鸡|一群鸡|鸡群|养鸡|鸡叽叽喳喳|鸡都|鸡呢|鸡……|鸡，|鸡。|鸡、)"),
        ("鸭子", r"(?:鸭子|鸭群|养鸭|新送来的鸭子|新送来的鸭|鸭，|鸭。)"),
        ("狗", r"(?:一只狗|几只狗|小狗|大狗|狗叫|牵着狗|带着狗|狗在|狗跑)"),
        ("猫", r"(?:一只猫|几只猫|小猫|猫叫|抱着猫|猫在|猫跑)"),
        ("鹅", r"(?:一只鹅|几只鹅|鹅群|养鹅)"),
        ("猪", r"(?:一头猪|几头猪|猪圈|养猪)"),
        ("牛", r"(?:一头牛|几头牛|牛棚|牛羊|养牛)"),
        ("羊", r"(?:一只羊|几只羊|羊群|牛羊|养羊)"),
    ]
    for animal, pattern in animal_patterns:
        if re.search(pattern, text):
            animals.append(animal)
    return ordered_unique(animals)


def has_live_animal(text: str) -> bool:
    cleaned = text.replace("阿猫阿狗", "")
    return bool(re.search(r"几只鸡|一只鸡|这批鸡|鸡叽叽喳喳|鸭子|养鸭|一只狗|几只狗|小狗|一只猫|几只猫|鹅群|猪圈|牛羊|羊群", cleaned))


def infer_space_type(location: str, text: str) -> str:
    combined = location + "\n" + text
    if any(keyword in location for keyword in ("手机屏幕", "屏幕", "一组镜头", "蒙太奇", "闪回", "转场")):
        return "插入/特殊画面"
    if any(keyword in location for keyword in ("老宅", "家", "房间", "屋")):
        return "实景/可搭景"
    if any(keyword in combined for keyword in ("医院", "诊所", "医馆", "学校", "公司", "办公室", "会议室", "酒店", "民宿", "餐馆", "饭店")):
        return "实景改造/可租景"
    if any(keyword in combined for keyword in ("村", "桥", "路口", "山", "田", "街", "广场", "院")):
        return "实景外景"
    return "待定"


def infer_atmosphere(scene: dict) -> str:
    location = scene["location"]
    text = scene["body_text"]
    combined = location + "\n" + text
    if any(keyword in combined for keyword in ("吵", "闹", "围", "争", "骂", "冲突", "打", "砸")):
        return "冲突紧张、人物密集"
    if any(keyword in combined for keyword in ("病", "救", "药", "医院", "诊所", "医馆")):
        return "医疗紧迫、克制冷静"
    if any(keyword in combined for keyword in ("修缮", "翻新", "施工", "桥")):
        return "乡村建设、旧新对比"
    if any(keyword in combined for keyword in ("直播", "游客", "民宿", "文旅", "开业")):
        return "经营热闹、文旅烟火气"
    if any(keyword in combined for keyword in ("老宅", "旧", "祖", "回忆", "闪回")):
        return "年代感、旧物记忆"
    if "会议" in combined:
        return "基层会议、公共秩序"
    return "现实生活质感"


def infer_space_state(scene: dict) -> str:
    text = scene["body_text"]
    location = scene["location"]
    combined = location + "\n" + text
    states = []
    if re.search(r"破|旧|脏|乱|荒|废", combined):
        states.append("破旧/待整理")
    if re.search(r"修缮|翻新|施工|改造|重建", combined):
        states.append("修缮/改造中")
    if re.search(r"砸|摔|碎|破坏|血|水泼|倒", combined):
        states.append("损坏需接戏")
    if re.search(r"会议|开会|讨论", combined):
        states.append("会议状态")
    if re.search(r"直播|游客|开业|热闹|围观", combined):
        states.append("人群经营状态")
    if re.search(r"夜|灯|油灯|蜡烛", combined) or scene["day"] == "夜":
        states.append("夜景用光状态")
    return "、".join(ordered_unique(states)) or "日常状态"


def location_set_note(location: str, text: str) -> str:
    rule = match_location_rule(location)
    if rule:
        return rule["wall"]
    if any(keyword in location for keyword in INSERT_KEYWORDS):
        return "准备插入画面内容或可拍屏幕方案"
    if "外" in text:
        return "确认外景空间边界、背景可控物和遮挡"
    return "按场景身份补齐空间标识和基础陈设"


def location_dressing_note(location: str) -> str:
    rule = match_location_rule(location)
    if rule:
        return "；".join([rule["furniture"], rule["soft"]])
    if any(keyword in location for keyword in INSERT_KEYWORDS):
        return "画面图文、界面素材、道具台面按剧本制作/确认"
    return "家具、生活痕迹、可见文字按剧情补齐"


def extract_scene_props(scene: dict) -> list[dict]:
    text = scene["location"] + "\n" + scene["people_raw"] + "\n" + scene["body_text"]
    props = []
    for rule in PROP_RULES:
        if rule["name"] == "活物/动物":
            for animal in extract_live_animals(scene):
                props.append(
                    {
                        "name": animal,
                        "keywords": (animal,),
                        "category": "特殊道具/活物",
                        "purpose": f"{animal}活物调度",
                    }
                )
            continue
        if any(keyword in text for keyword in rule["keywords"]):
            props.append(rule)
    return props


PERSONAL_PROP_LABELS = {
    "手机/屏幕画面": "手机",
    "行李/包袋": "包袋/行李",
    "车辆/车钥匙": "车辆/车钥匙",
}


def line_mentions_owner(line: str, person: str, keywords: tuple[str, ...]) -> bool:
    person_pattern = re.escape(person)
    prop_pattern = "|".join(re.escape(keyword) for keyword in keywords)
    return bool(
        re.search(person_pattern + rf".{{0,12}}(?:{prop_pattern})", line)
        or re.search(rf"(?:{prop_pattern}).{{0,12}}" + person_pattern, line)
        or re.search(person_pattern + r".{0,12}(接起|拨通|拿出|掏出|背着|拎着|打开|开车|上车|下车)", line)
    )


def infer_prop_owners(scene: dict, prop: dict) -> list[tuple[str, str]]:
    base = prop["name"]
    text = scene["location"] + "\n" + scene["people_raw"] + "\n" + scene["body_text"]
    people = named_people(scene)
    if base not in PERSONAL_PROP_LABELS:
        return [("共用/场景", "场景/陈设")]

    if base == "手机/屏幕画面" and any(keyword in scene["location"] for keyword in ("手机屏幕", "屏幕")):
        return [("画面/屏幕内容", "屏幕/图文")]

    owners = []
    for person in people:
        for line in scene.get("body", []):
            if line_mentions_owner(line, person, prop["keywords"]):
                owners.append((person, "演员随身"))
                break

    if owners:
        return ordered_unique_pairs(owners)

    if len(people) == 1:
        if base == "手机/屏幕画面" and re.search(r"手机|电话|通话|接起|拨通", text):
            return [(people[0], "演员随身")]
        if base == "行李/包袋" and re.search(r"行李|箱子|背包|提包|包", text):
            return [(people[0], "演员随身")]

    if base == "手机/屏幕画面" and "直播" in text:
        return [("直播/拍摄", "电子设备")]
    if base == "车辆/车钥匙":
        return [("交通道具/待定", "交通道具")]
    return [("未明确归属", "需确认归属")]


def ordered_unique_pairs(items: list[tuple[str, str]]) -> list[tuple[str, str]]:
    seen = OrderedDict()
    for owner, scope in items:
        if owner:
            seen[(owner, scope)] = True
    return list(seen.keys())


def prop_instance_name(base: str, owner: str, scope: str) -> str:
    label = PERSONAL_PROP_LABELS.get(base, base)
    if owner in {"共用/场景"}:
        return base
    if owner == "画面/屏幕内容":
        return "手机屏幕画面" if base == "手机/屏幕画面" else f"{label}画面"
    if owner == "直播/拍摄":
        return f"直播/拍摄{label}"
    if owner == "交通道具/待定":
        return label
    if owner in {"待定", "未明确归属"}:
        return f"未明确归属{label}（需确认）"
    if scope == "演员随身":
        return f"{owner}{label}"
    return f"{owner}{label}"


def extract_prop_instances(scene: dict) -> list[dict]:
    instances = []
    for prop in extract_scene_props(scene):
        for owner, scope in infer_prop_owners(scene, prop):
            instances.append(
                {
                    "name": prop_instance_name(prop["name"], owner, scope),
                    "base": prop["name"],
                    "category": prop["category"],
                    "purpose": prop["purpose"],
                    "owner": owner,
                    "scope": scope,
                }
            )
    unique = OrderedDict()
    for item in instances:
        key = (item["name"], item["owner"], item["scope"], item["category"])
        unique[key] = item
    return list(unique.values())


def scene_prop_text(scene: dict) -> str:
    props = [prop["name"] for prop in extract_prop_instances(scene)]
    return "、".join(ordered_unique(props)) or "无明确戏用道具"


def infer_special(scene: dict) -> str:
    text = scene["location"] + "\n" + scene["body_text"]
    notes = []
    if any(keyword in text for keyword in ("手机屏幕", "屏幕", "短信", "微信", "直播", "视频")):
        notes.append("屏幕/界面内容需制作审核")
    if any(keyword in text for keyword in ("字幕", "画外音", "VO", "旁白")):
        notes.append("字幕/画外信息需对齐")
    if any(keyword in text for keyword in ("闪回", "回忆", "闪现", "蒙太奇", "一组镜头")):
        notes.append("闪回/蒙太奇画面风格统一")
    if any(keyword in text for keyword in ("砸", "摔碎", "打碎", "破", "血", "水泼", "火", "刀")):
        notes.append("动作损坏道具需备份")
    if any(keyword in text for keyword in ("签字", "合同", "协议", "文件", "证书", "台账", "登记")):
        notes.append("图文道具需提前出版面")
    if scene.get("warning"):
        notes.append(scene["warning"])
    return "；".join(ordered_unique(notes)) or "无"


def infer_continuity(scene: dict, location_occurrences: dict[str, list[str]]) -> str:
    text = scene["body_text"]
    loc = normalize_location(scene["location"])
    notes = []
    if len(location_occurrences.get(loc, [])) > 1:
        notes.append("同场景多次出现，陈设位置需接戏")
    if re.search(r"修缮|翻新|施工|改造|重建|破|砸|碎|血|水泼", text):
        notes.append("空间/道具状态变化需拍照记录")
    if re.search(r"现金|文件|合同|手机|药|照片|牌匾|招牌", text):
        notes.append("关键道具版本和位置需统一")
    return "；".join(ordered_unique(notes)) or "常规接戏"


def infer_risk(scene: dict) -> str:
    text = scene["location"] + "\n" + scene["body_text"]
    risks = []
    if re.search(r"合同|协议|证书|病历|处方|台账|名单|手机屏幕|直播|字幕", text):
        risks.append("图文内容需审核")
    if re.search(r"医院|诊所|医馆|药|病历|诊断|输液", text):
        risks.append("医疗专业性需核对")
    if re.search(r"砸|摔|打|刀|血|火|水泼|破", text):
        risks.append("动作/损坏安全与备份")
    if re.search(r"群众|众人|围观|游客|村民|会议", text):
        risks.append("人群空间和陈设避穿帮")
    if re.search(r"修缮|施工|桥|老宅|旧", text):
        risks.append("旧化/施工状态连续性")
    return "；".join(ordered_unique(risks)) or "低"


def prop_quantity(category: str, scenes: list[str]) -> str:
    if "消耗" in category or "损坏" in category or "危险" in category:
        return "按动作和重拍备多套"
    if "图文" in category or "屏幕" in category:
        return "主版+备份版"
    if len(scenes) > 3:
        return "主道具+备份"
    return "1套起，按调度备份"


def prop_source(category: str) -> str:
    if "图文" in category:
        return "美术设计制作"
    if "专业" in category:
        return "租赁/采购+专业核对"
    if "损坏" in category or "危险" in category:
        return "安全道具制作/多套备份"
    if "消耗" in category:
        return "采购+连续性备量"
    return "采购/租赁"


def build_location_occurrences(scenes: list[dict]) -> dict[str, list[str]]:
    occurrences: dict[str, list[str]] = defaultdict(list)
    for scene in scenes:
        occurrences[normalize_location(scene["location"])].append(scene["scene"])
    return occurrences


def aggregate_props(scenes: list[dict]) -> list[dict]:
    props: dict[str, dict] = {}
    for scene in scenes:
        for prop in extract_prop_instances(scene):
            name = prop["name"]
            entry = props.setdefault(
                name,
                {
                    "name": name,
                    "base": prop["base"],
                    "category": prop["category"],
                    "purpose": prop["purpose"],
                    "owner": prop["owner"],
                    "scope": prop["scope"],
                    "scenes": [],
                    "locations": [],
                },
            )
            entry["scenes"].append(scene["scene"])
            entry["locations"].append(normalize_location(scene["location"]))
    return list(props.values())


def aggregate_locations(scenes: list[dict]) -> list[dict]:
    grouped: dict[str, dict] = {}
    for scene in scenes:
        loc = normalize_location(scene["location"])
        main_set, subspace = split_location_for_scouting(scene["location"], scene["inside_outside"])
        entry = grouped.setdefault(
            loc,
            {
                "location": loc,
                "main_set": main_set,
                "subspace": subspace,
                "scenes": [],
                "day_io": [],
                "texts": [],
                "raw_locations": [],
            },
        )
        entry["scenes"].append(scene["scene"])
        entry["day_io"].append(f"{scene['day']}/{scene['inside_outside']}")
        entry["texts"].append(scene["body_text"])
        entry["raw_locations"].append(scene["location"])
    return list(grouped.values())


def aggregate_scouting_spaces(scenes: list[dict]) -> list[dict]:
    grouped: OrderedDict[tuple[str, str], dict] = OrderedDict()
    main_subspaces: dict[str, list[str]] = defaultdict(list)
    for scene in scenes:
        main_set, subspace = split_location_for_scouting(scene["location"], scene["inside_outside"])
        main_subspaces[main_set].append(subspace)
        key = (main_set, subspace)
        entry = grouped.setdefault(
            key,
            {
                "main_set": main_set,
                "subspace": subspace,
                "scenes": [],
                "day_io": [],
                "texts": [],
                "locations": [],
                "props": [],
            },
        )
        entry["scenes"].append(scene["scene"])
        entry["day_io"].append(f"{scene['day']}/{scene['inside_outside']}")
        entry["texts"].append(scene["body_text"])
        entry["locations"].append(normalize_location(scene["location"]))
        entry["props"].extend([prop["name"] for prop in extract_prop_instances(scene)])

    result = []
    for entry in grouped.values():
        entry["all_subspaces"] = ordered_unique(main_subspaces[entry["main_set"]])
        result.append(entry)
    return result


def location_priority(entry: dict) -> str:
    text = "\n".join(entry["texts"]) + entry["location"]
    if len(entry["scenes"]) >= 5 or re.search(r"桥|老宅|医院|医馆|直播|施工|修缮|砸|破|合同|手机屏幕", text):
        return "高"
    if len(entry["scenes"]) >= 2:
        return "中"
    return "低"


def location_state_change(entry: dict) -> str:
    text = "\n".join(entry["texts"]) + entry["location"]
    states = []
    if re.search(r"旧|破|脏|乱|荒", text):
        states.append("旧/破状态")
    if re.search(r"修缮|翻新|施工|改造|重建", text):
        states.append("修缮/翻新状态")
    if re.search(r"砸|摔碎|打碎|血|水泼|火", text):
        states.append("损坏/污染状态")
    if len(set(entry["day_io"])) > 1:
        states.append("日夜/内外多状态")
    return "、".join(ordered_unique(states)) or "常规状态"


def reuse_advice(entry: dict) -> str:
    if len(entry["scenes"]) >= 4:
        return "建议固定主场景，按时间线记录陈设变化"
    if location_priority(entry) == "高":
        return "提前定景，预留改景和复位时间"
    return "可按同类空间复用或轻改"


def scouting_condition(entry: dict) -> str:
    subspaces = "、".join(entry.get("all_subspaces", []))
    text = "\n".join(entry["texts"]) + entry["main_set"] + entry["subspace"]
    space_type = infer_space_type(entry["main_set"], text)
    conditions = [space_type]
    if subspaces:
        conditions.append(f"需覆盖：{subspaces}")
    if any(keyword in text for keyword in ("群众", "众人", "围观", "会议", "游客")):
        conditions.append("需容纳多人调度")
    if any(keyword in text for keyword in ("修缮", "施工", "砸", "破", "旧")):
        conditions.append("可做旧化/施工/损坏状态")
    if any(keyword in text for keyword in ("夜", "灯", "油灯")):
        conditions.append("夜景和实用灯位可控")
    return "；".join(ordered_unique(conditions))


def scouting_dressing(entry: dict) -> str:
    notes = []
    for location in ordered_unique(entry["locations"]):
        notes.append(location_dressing_note(location))
    return "；".join(ordered_unique(notes)) or "按空间身份补齐基础陈设"


def scouting_set_note(entry: dict) -> str:
    notes = []
    for location in ordered_unique(entry["locations"]):
        notes.append(location_set_note(location, "\n".join(entry["texts"])))
    return "；".join(ordered_unique(notes)) or "按主场景身份确定空间风格"


def scouting_reuse_advice(entry: dict) -> str:
    if len(entry["scenes"]) >= 4 or len(entry.get("all_subspaces", [])) >= 3:
        return "优先找同一实景覆盖多子空间，统一美术风格"
    if len(entry.get("all_subspaces", [])) >= 2:
        return "建议同一主场景内轻改或相邻空间完成"
    return "可独立找景或与同类空间复用"


def build_personal_props(props: list[dict]) -> list[dict]:
    grouped: OrderedDict[tuple[str, str], dict] = OrderedDict()
    for prop in props:
        if prop.get("scope") != "演员随身":
            continue
        key = (prop["owner"], prop["name"])
        entry = grouped.setdefault(
            key,
            {
                "owner": prop["owner"],
                "name": prop["name"],
                "scenes": [],
                "locations": [],
            },
        )
        entry["scenes"].extend(prop["scenes"])
        entry["locations"].extend(prop["locations"])
    return list(grouped.values())


def build_continuity_risks(scenes: list[dict], locations: list[dict], props: list[dict]) -> list[list[str]]:
    rows = []
    for entry in locations:
        if location_priority(entry) in {"高", "中"}:
            rows.append(
                [
                    entry["location"],
                    "空间状态",
                    "、".join(entry["scenes"]),
                    location_state_change(entry),
                    reuse_advice(entry),
                    location_priority(entry),
                ]
            )
    for prop in props:
        if any(keyword in prop["category"] for keyword in ("图文", "屏幕", "损坏", "危险", "专业", "消耗")) or len(prop["scenes"]) >= 3:
            rows.append(
                [
                    prop["name"],
                    f"{prop['category']} / {prop.get('scope', '')}",
                    "、".join(ordered_unique(prop["scenes"])),
                    f"{prop['purpose']}；归属/版本/状态需统一",
                    "建立道具照片、版面文件和场记交接记录",
                    "高" if any(k in prop["category"] for k in ("图文", "损坏", "危险", "专业")) or prop.get("scope") == "演员随身" else "中",
                ]
            )
    for scene in scenes:
        risk = infer_risk(scene)
        if risk != "低" and any(keyword in risk for keyword in ("动作", "医疗", "图文")):
            rows.append(
                [
                    f"{scene['scene']} {normalize_location(scene['location'])}",
                    "分场风险",
                    scene["scene"],
                    risk,
                    "美术、道具、场记和相关部门提前确认方案",
                    "高" if "动作" in risk or "医疗" in risk else "中",
                ]
            )
    return rows


def title_from_path(path: Path) -> str:
    match = re.search(r"《([^》]+)》", path.stem)
    if match:
        return match.group(1)
    return path.stem


def style_sheet(ws, widths: dict[int, float], header_rows: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="263A4A")
    sub_fill = PatternFill("solid", fgColor="E7EEF3")
    alt_fill = PatternFill("solid", fgColor="F6F8FA")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="C9D3DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=10)
    body_font = Font(name="Microsoft YaHei", color="222222", size=10)

    ws.freeze_panes = f"A{header_rows + 1}"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row >= header_rows:
        ws.auto_filter.ref = f"A{header_rows}:{get_column_letter(max_col)}{max_row}"
    for row in range(1, max_row + 1):
        fill = alt_fill if row > header_rows and row % 2 == 0 else white_fill
        if row <= header_rows:
            fill = header_fill if row == header_rows else sub_fill
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = border
            cell.font = header_font if row == header_rows else body_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 38 if row > header_rows else 42

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    for row in range(header_rows + 1, max_row + 1):
        for col in range(1, max_col + 1):
            if ws.column_dimensions[get_column_letter(col)].width >= 16:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def add_overview_sheet(wb: Workbook, title: str, scenes: list[dict], locations: list[dict], props: list[dict]) -> None:
    ws = wb.create_sheet("美术风格总览", 0)
    text_all = "\n".join(scene["location"] + "\n" + scene["body_text"] for scene in scenes)
    tones = ordered_unique([infer_atmosphere(scene) for scene in scenes])[:6]
    high_locations = [entry["location"] for entry in locations if location_priority(entry) == "高"][:10]
    graphic_props = [prop["name"] for prop in props if "图文" in prop["category"] or "屏幕" in prop["category"]]
    risk_items = []
    for scene in scenes:
        risk = infer_risk(scene)
        if risk != "低":
            risk_items.extend(risk.split("；"))
    rows = [
        ("项目", title),
        ("场次数", str(len(scenes))),
        ("主要空间", "、".join(high_locations) or "待定"),
        ("视觉基调", "；".join(tones) or "现实生活质感"),
        ("色彩方向", "乡村自然色、旧木色、基层办公灰白、经营空间暖色；具体按导演摄影方案确认"),
        ("材质方向", "旧墙面、木质家具、纸质图文、生活杂物、施工修缮质感"),
        ("重复出现道具", "、".join(prop["name"] for prop in props[:12]) or "待定"),
        ("图文重点", "、".join(ordered_unique(graphic_props)) or "暂无明确图文道具"),
        ("高风险提示", "；".join(ordered_unique(risk_items)) or "常规"),
        ("参考检索词", build_reference_prompts(title, text_all)),
    ]
    ws.append(["项目", "内容"])
    for row in rows:
        ws.append(row)
    style_sheet(ws, {1: 18, 2: 90})
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 48


def build_reference_prompts(title: str, text: str) -> str:
    prompts = [f"{title} 乡村现实主义 美术参考"]
    if "桥" in text:
        prompts.append("乡村石桥修缮 文旅改造 视觉参考")
    if re.search(r"村委会|会议", text):
        prompts.append("村委会会议室 基层办公 陈设参考")
    if re.search(r"民宿|游客|直播", text):
        prompts.append("乡村民宿 直播带货 美术参考")
    if re.search(r"老宅|旧", text):
        prompts.append("乡村老宅 旧化 生活痕迹 参考")
    return "；".join(ordered_unique(prompts))


def join_items(items: list[str], fallback: str = "") -> str:
    return "、".join(ordered_unique(items)) or fallback


def is_special_prop(prop: dict) -> bool:
    name = prop.get("name", "")
    category = prop.get("category", "")
    if name == "手机屏幕画面" or "直播" in name:
        return True
    text = f"{name}{category}"
    return bool(
        re.search(
            r"车辆|车钥匙|交通|活物|动物|鸡鸭|鸡群|鸭群|狗|猫|鹅|猪|牛羊|羊群|危险|损坏|破损|刀|血|火|水泼|医疗|专业|施工",
            text,
        )
    )


def prop_buckets_for_scene(scene: dict) -> dict[str, list[str]]:
    buckets = {"set": [], "action": [], "special": []}
    for prop in extract_prop_instances(scene):
        if is_special_prop(prop):
            buckets["special"].append(prop["name"])
            continue
        if prop.get("scope") == "场景/陈设" and (
            "陈设" in prop["category"] or "置景" in prop["category"] or prop["base"] in {"会议用品", "灯具/油灯", "农具/乡村用具"}
        ):
            buckets["set"].append(prop["name"])
        else:
            buckets["action"].append(prop["name"])
    return {key: ordered_unique(value) for key, value in buckets.items()}


def basic_set_dressing(main_set: str, subspace: str, location: str) -> list[str]:
    text = f"{main_set}{subspace}{location}"
    if any(keyword in text for keyword in ("村委会", "会议室")):
        return ["会议桌椅", "公告栏/制度牌", "文件资料", "茶杯/水杯"]
    if any(keyword in text for keyword in ("医院", "诊所", "医馆", "病房")):
        return ["诊疗桌椅", "药柜/药品陈列", "病历/处方资料", "医疗标识"]
    if any(keyword in text for keyword in ("厨房", "灶")):
        return ["灶台", "锅碗瓢盆", "餐具", "生活杂物"]
    if any(keyword in text for keyword in ("客厅", "厅")):
        return ["沙发/椅子", "茶几", "电视柜/柜子", "生活摆件"]
    if any(keyword in text for keyword in ("卧室", "房间", "屋", "小屋", "宿舍")):
        return ["床铺", "柜子", "桌椅", "生活用品"]
    if any(keyword in text for keyword in ("院子", "院内", "院外", "前院", "后院")):
        return ["院落杂物", "农具", "水桶/盆", "晾晒物"]
    if any(keyword in text for keyword in ("门口", "门外", "门前")):
        return ["门牌/门口标识", "门前杂物", "外立面基础陈设"]
    if "家" in text and any(keyword in text for keyword in ("外", "整体外景")):
        return ["房屋外立面", "门口生活杂物", "院落基础陈设"]
    if "家" in text:
        return ["桌椅/柜子", "生活杂物", "家庭照片/墙面痕迹"]
    if any(keyword in text for keyword in ("路口", "村口", "小路", "桥", "街", "广场")):
        return ["路牌/村名标识", "墙面标语", "路边杂物", "实景基础环境"]
    if any(keyword in text for keyword in INSERT_KEYWORDS):
        return ["插入画面所需背景/台面", "画面图文素材"]
    return ["按实景基础陈设"]


def scene_record(scene: dict, order: int) -> dict:
    main_set, subspace = split_location_for_scouting(scene["location"], scene["inside_outside"])
    buckets = prop_buckets_for_scene(scene)
    set_props = buckets["set"] or basic_set_dressing(main_set, subspace, scene["location"])
    return {
        "order": order,
        "scene": scene["scene"],
        "source_page": scene.get("source_page", ""),
        "main_set": main_set,
        "subspace": subspace,
        "location": scene["location"],
        "day": scene["day"],
        "inside_outside": scene["inside_outside"],
        "people": people_and_groups(scene),
        "summary": scene_summary(scene),
        "set_props": set_props,
        "action_props": buckets["action"],
        "special_props": buckets["special"],
        "note": scene.get("warning", ""),
        "body_text": scene["body_text"],
    }


def aggregate_scene_table(records: list[dict]) -> list[dict]:
    grouped: OrderedDict[str, dict] = OrderedDict()
    for record in records:
        entry = grouped.setdefault(
            record["main_set"],
            {
                "main_set": record["main_set"],
                "subspaces": [],
                "scenes": [],
                "set_props": [],
                "action_props": [],
                "special_props": [],
                "notes": [],
            },
        )
        entry["subspaces"].append(record["subspace"])
        entry["scenes"].append(record["scene"])
        entry["set_props"].extend(record["set_props"])
        entry["action_props"].extend(record["action_props"])
        entry["special_props"].extend(record["special_props"])
        if record["note"]:
            entry["notes"].append(record["note"])
    return list(grouped.values())


def aggregate_execution_table(records: list[dict]) -> list[dict]:
    grouped: OrderedDict[tuple[str, str], dict] = OrderedDict()
    for record in records:
        key = (record["main_set"], record["subspace"])
        entry = grouped.setdefault(
            key,
            {
                "main_set": record["main_set"],
                "subspace": record["subspace"],
                "scenes": [],
                "locations": [],
                "texts": [],
                "set_props": [],
                "action_props": [],
                "special_props": [],
            },
        )
        entry["scenes"].append(record["scene"])
        entry["locations"].append(normalize_location(record["location"]))
        entry["texts"].append(record["body_text"])
        entry["set_props"].extend(record["set_props"])
        entry["action_props"].extend(record["action_props"])
        entry["special_props"].extend(record["special_props"])
    return list(grouped.values())


def setup_content_for_entry(entry: dict) -> str:
    notes = []
    text = "\n".join(entry.get("texts", []))
    for location in ordered_unique(entry.get("locations", [entry.get("main_set", "")])):
        notes.append(location_set_note(location, text))
        notes.append(location_dressing_note(location))
    return "；".join(ordered_unique(notes)) or "按主场景和次场景完成基础布置"


def purchase_items(entry: dict) -> list[str]:
    items = []
    all_props = entry.get("set_props", []) + entry.get("action_props", []) + entry.get("special_props", [])
    for item in all_props:
        if any(keyword in item for keyword in ("餐饮", "现金", "红包", "银行卡", "包袋", "灯具", "油灯", "农具", "会议用品")):
            items.append(item)
    return ordered_unique(items)


def fabrication_items(entry: dict) -> list[str]:
    items = []
    all_props = entry.get("set_props", []) + entry.get("action_props", []) + entry.get("special_props", [])
    for item in all_props:
        if any(keyword in item for keyword in ("文件", "合同", "台账", "屏幕", "牌匾", "招牌", "门牌", "标识", "横幅", "损坏", "破损")):
            items.append(item)
    return ordered_unique(items)


def rental_items(entry: dict) -> list[str]:
    items = []
    all_props = entry.get("set_props", []) + entry.get("action_props", []) + entry.get("special_props", [])
    for item in all_props:
        if any(keyword in item for keyword in ("车辆", "车钥匙", "医药", "直播", "施工", "活物", "动物")):
            items.append(item)
    return ordered_unique(items)


def apply_special_red(ws, header_name: str) -> None:
    target_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value == header_name:
            target_col = col
            break
    if target_col is None:
        return
    red_font = Font(name="Microsoft YaHei", color="C00000", bold=True, size=10)
    red_fill = PatternFill("solid", fgColor="FCE4D6")
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, target_col)
        if cell.value not in (None, "", "无"):
            cell.font = red_font
            cell.fill = red_fill


def merge_repeated_values(ws, header_name: str, start_row: int = 2) -> None:
    target_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value == header_name:
            target_col = col
            break
    if target_col is None:
        return

    row = start_row
    while row <= ws.max_row:
        value = ws.cell(row, target_col).value
        end = row
        while end + 1 <= ws.max_row and ws.cell(end + 1, target_col).value == value:
            end += 1
        if value not in (None, "") and end > row:
            ws.merge_cells(start_row=row, start_column=target_col, end_row=end, end_column=target_col)
            cell = ws.cell(row, target_col)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row = end + 1


def build_workbook(source: Path, scenes: list[dict], output: Path, title: str) -> None:
    if not scenes:
        raise RuntimeError(f"No scenes parsed from {source}")

    records = [scene_record(scene, order) for order, scene in enumerate(scenes, start=1)]
    grouped_scene_rows = aggregate_scene_table(records)
    execution_rows = aggregate_execution_table(records)

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("顺场表")
    ws.append(
        [
            "序号",
            "场次",
            "主场景",
            "次场景",
            "剧本中场景",
            "日/夜",
            "内/外",
            "人物/群演",
            "拍摄内容",
            "陈设道具",
            "戏用道具",
            "车辆/特殊道具",
            "备注",
        ]
    )
    for record in records:
        ws.append(
            [
                record["order"],
                record["scene"],
                record["main_set"],
                record["subspace"],
                record["location"],
                record["day"],
                record["inside_outside"],
                record["people"],
                record["summary"],
                join_items(record["set_props"], "按实景基础陈设"),
                join_items(record["action_props"], "无明确戏用道具"),
                join_items(record["special_props"], ""),
                record["note"],
            ]
        )
    style_sheet(ws, {1: 7, 2: 10, 3: 18, 4: 16, 5: 22, 6: 6, 7: 6, 8: 18, 9: 28, 10: 28, 11: 28, 12: 24, 13: 20})
    apply_special_red(ws, "车辆/特殊道具")

    ws = wb.create_sheet("分场表")
    ws.append(["主场景", "次场景", "场次", "剧本中场景", "日/夜", "内/外", "人物/群演", "拍摄内容", "陈设道具", "戏用道具", "车辆/特殊道具", "备注"])
    for record in sorted(records, key=lambda item: (item["main_set"], item["subspace"], item["order"])):
        ws.append(
            [
                record["main_set"],
                record["subspace"],
                record["scene"],
                record["location"],
                record["day"],
                record["inside_outside"],
                record["people"],
                record["summary"],
                join_items(record["set_props"], "按实景基础陈设"),
                join_items(record["action_props"], "无明确戏用道具"),
                join_items(record["special_props"], ""),
                record["note"],
            ]
        )
    style_sheet(ws, {1: 18, 2: 16, 3: 10, 4: 22, 5: 6, 6: 6, 7: 18, 8: 28, 9: 28, 10: 28, 11: 24, 12: 20})
    apply_special_red(ws, "车辆/特殊道具")
    merge_repeated_values(ws, "主场景")

    ws = wb.create_sheet("场景表")
    ws.append(["主场景", "包含次场景", "涉及场次", "陈设道具", "戏用道具", "车辆/特殊道具", "备注"])
    for entry in grouped_scene_rows:
        ws.append(
            [
                entry["main_set"],
                join_items(entry["subspaces"]),
                join_items(entry["scenes"]),
                join_items(entry["set_props"], "按实景基础陈设"),
                join_items(entry["action_props"], "无明确戏用道具"),
                join_items(entry["special_props"], ""),
                join_items(entry["notes"]),
            ]
        )
    style_sheet(ws, {1: 22, 2: 34, 3: 48, 4: 36, 5: 36, 6: 30, 7: 24})
    apply_special_red(ws, "车辆/特殊道具")

    ws = wb.create_sheet("置景道具执行表")
    ws.append(["主场景", "次场景", "需要布置内容", "陈设道具", "戏用道具", "需采购", "需制作", "需租赁/借用", "车辆/特殊道具", "备注"])
    for entry in sorted(execution_rows, key=lambda item: (item["main_set"], item["subspace"])):
        ws.append(
            [
                entry["main_set"],
                entry["subspace"],
                setup_content_for_entry(entry),
                join_items(entry["set_props"], "按实景基础陈设"),
                join_items(entry["action_props"], "无明确戏用道具"),
                join_items(purchase_items(entry)),
                join_items(fabrication_items(entry)),
                join_items(rental_items(entry)),
                join_items(entry["special_props"], ""),
                join_items(entry["scenes"]),
            ]
        )
    style_sheet(ws, {1: 22, 2: 16, 3: 42, 4: 34, 5: 34, 6: 28, 7: 28, 8: 28, 9: 28, 10: 36})
    apply_special_red(ws, "车辆/特殊道具")
    merge_repeated_values(ws, "主场景")

    wb.save(output)


def dump_json(path: Path, all_scenes: dict[str, list[dict]]) -> None:
    serializable = {}
    for key, scenes in all_scenes.items():
        serializable[key] = [
            {k: v for k, v in scene.items() if k not in {"body_text"}}
            for scene in scenes
        ]
    path.write_text(json.dumps(serializable, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build art department breakdown Excel files from scripts.")
    parser.add_argument("scripts", nargs="+", type=Path)
    parser.add_argument("--out-dir", type=Path, default=Path.cwd())
    parser.add_argument("--title", help="Override title/output name when processing a single script")
    parser.add_argument("--dump-json", type=Path)
    args = parser.parse_args()

    if args.title and len(args.scripts) != 1:
        raise SystemExit("--title can only be used with one script")

    args.out_dir.mkdir(parents=True, exist_ok=True)
    all_scenes: dict[str, list[dict]] = {}
    for source in args.scripts:
        scenes = parse_script(source)
        all_scenes[source.name] = scenes
        title = args.title or title_from_path(source)
        output = args.out_dir / f"美术拆解-{title}.xlsx"
        build_workbook(source, scenes, output, title)
        warnings = [f"{scene['scene']}:{scene['warning']}" for scene in scenes if scene.get("warning")]
        print(f"{source.name}: {len(scenes)} scenes -> {output}")
        if warnings:
            print("warnings:", "；".join(warnings))

    if args.dump_json:
        dump_json(args.dump_json, all_scenes)
        print(f"scene dump -> {args.dump_json}")


if __name__ == "__main__":
    main()
