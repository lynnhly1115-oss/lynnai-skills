#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from collections import OrderedDict
from pathlib import Path

from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import pdfplumber
except ImportError:  # pragma: no cover - dependency depends on the active runtime
    pdfplumber = None


CN_EPISODES = {
    "一": 1,
    "二": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
    "十": 10,
    "十一": 11,
    "十二": 12,
    "十三": 13,
    "十四": 14,
    "十五": 15,
    "十六": 16,
    "十七": 17,
    "十八": 18,
    "十九": 19,
    "二十": 20,
    "二十一": 21,
    "二十二": 22,
    "二十三": 23,
    "二十四": 24,
    "二十五": 25,
    "二十六": 26,
    "二十七": 27,
    "二十八": 28,
    "二十九": 29,
    "三十": 30,
}

GROUP_HINTS = (
    "若干",
    "众人",
    "员工",
    "职员",
    "应聘者",
    "高层",
    "保镖",
    "村民",
    "群众",
    "游客",
    "干部",
    "工作人员",
    "代表",
    "老村民",
    "股东",
    "参会人",
    "客人",
    "夜枭成员",
    "保安",
)

DAY_RE = r"日|夜|清晨|上午|下午|傍晚|深夜|凌晨"
IO_RE = r"内外|内|外"

PROP_KEYWORDS = [
    ("轮椅", "轮椅"),
    ("防爆盾", "防爆盾"),
    ("电棍", "电棍"),
    ("子弹", "子弹/枪械效果"),
    ("狙击手", "枪械效果"),
    ("婴儿车", "婴儿车"),
    ("奶瓶", "奶瓶"),
    ("纸尿裤", "纸尿裤"),
    ("襁褓", "襁褓"),
    ("书信", "书信"),
    ("推荐信", "推荐信"),
    ("门把手", "门把手"),
    ("清单", "清单"),
    ("运动衣", "运动衣"),
    ("菜刀", "菜刀"),
    ("鞭子", "鞭子"),
    ("遥控器", "遥控器"),
    ("炸弹", "炸弹道具"),
    ("棍棒", "棍棒"),
    ("竞标方案", "竞标方案"),
    ("投影", "投影屏"),
    ("公章", "公章"),
    ("股权书", "股权书"),
    ("短刀", "短刀"),
    ("银针", "银针"),
    ("横幅", "横幅"),
    ("手机", "手机"),
    ("电话", "手机/电话"),
    ("平板", "平板"),
    ("电脑", "电脑"),
    ("文件夹", "文件夹"),
    ("文件", "文件"),
    ("合同", "合同"),
    ("协议", "协议"),
    ("信封", "信封"),
    ("现金", "现金"),
    ("两千块", "现金"),
    ("验孕棒", "验孕棒"),
    ("戒指", "戒指"),
    ("首饰盒", "首饰盒"),
    ("蛋糕", "蛋糕"),
    ("菜肴", "菜肴"),
    ("饮料", "饮料"),
    ("项链", "项链"),
    ("购物袋", "购物袋"),
    ("保温壶", "保温壶"),
    ("咖啡杯", "咖啡杯"),
    ("咖啡", "咖啡"),
    ("粉包", "粉包"),
    ("照片", "照片"),
    ("报告", "报告"),
    ("亲子鉴定", "亲子鉴定报告"),
    ("拐杖", "拐杖"),
    ("对讲机", "对讲机"),
    ("椅子", "椅子"),
    ("玻璃瓶", "玻璃瓶"),
    ("豪车", "豪车"),
    ("清洁工具", "清洁工具"),
    ("水桶", "水桶"),
    ("冷水", "冷水"),
    ("棉签", "棉签"),
    ("药品", "药品"),
    ("擦点药", "药品"),
    ("病床", "病床"),
    ("输液", "输液道具"),
    ("水杯", "水杯"),
    ("糕点", "糕点"),
    ("水果刀", "水果刀"),
    ("美工刀", "美工刀"),
    ("刀", "刀具"),
    ("花瓶", "花瓶"),
    ("财务报表", "财务报表"),
    ("催债函", "催债函"),
]


def cn_episode_to_int(value: str) -> int | None:
    if value.isdigit():
        return int(value)
    return CN_EPISODES.get(value)


def normalize_scene_heading(text: str) -> str:
    text = text.replace("　", " ").strip()
    text = re.sub(r"^[•·●○▪■□◆◇◦\-\s]+", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def clean_location(location: str) -> str:
    location = location.strip(" 　-=—")
    location = re.sub(r"\s*-\s*", "-", location)
    location = re.sub(r"\s+", " ", location)
    return location


def is_scene_heading(text: str) -> bool:
    text = normalize_scene_heading(text)
    if not text or any(mark in text for mark in "：:。？！；;"):
        return False
    explicit_day_first = rf"^\d+-\d+\s*(?:{DAY_RE})\s*(?:{IO_RE})\s*.+$"
    explicit_location_first = rf"^\d+-\d+\s*.+?(?:{DAY_RE})\s*(?:{IO_RE})$"
    generated_day_first = rf"^(?:{DAY_RE})\s*(?:{IO_RE})\s*.+$"
    generated_location_first = rf"^.+?(?:{DAY_RE})\s*(?:{IO_RE})$"
    return bool(
        re.match(explicit_day_first, text)
        or re.match(explicit_location_first, text)
        or re.match(generated_day_first, text)
        or re.match(generated_location_first, text)
    )


def parse_heading(text: str, episode: int | None, scene_count: dict[int, int], last_day: str) -> tuple[str, str, str, str, str] | None:
    text = normalize_scene_heading(text)
    explicit = re.match(
        rf"^(\d+)-(\d+)\s*(?:(?P<day>{DAY_RE})\s*)?(?P<io>{IO_RE})\s*(?P<location>.+)$",
        text,
    )
    if explicit:
        ep = int(explicit.group(1))
        no = int(explicit.group(2))
        scene_count[ep] = max(scene_count.get(ep, 0), no)
        day = explicit.group("day") or last_day or "日"
        warning = "" if explicit.group("day") else "day inferred"
        return f"{ep}-{no}", normalize_day(day), explicit.group("io"), clean_location(explicit.group("location")), warning

    explicit_location_first = re.match(
        rf"^(\d+)-(\d+)\s*(?P<location>.+?)\s*(?P<day>{DAY_RE})\s*(?P<io>{IO_RE})$",
        text,
    )
    if explicit_location_first:
        ep = int(explicit_location_first.group(1))
        no = int(explicit_location_first.group(2))
        scene_count[ep] = max(scene_count.get(ep, 0), no)
        return (
            f"{ep}-{no}",
            normalize_day(explicit_location_first.group("day")),
            explicit_location_first.group("io"),
            clean_location(explicit_location_first.group("location")),
            "",
        )

    generated = re.match(rf"^({DAY_RE})\s*({IO_RE})\s*(.+)$", text)
    if generated and episode:
        scene_count[episode] = scene_count.get(episode, 0) + 1
        return f"{episode}-{scene_count[episode]}", normalize_day(generated.group(1)), generated.group(2), clean_location(generated.group(3)), ""

    generated_location_first = re.match(
        rf"^(?P<location>.+?)\s*(?P<day>{DAY_RE})\s*(?P<io>{IO_RE})$",
        text,
    )
    if generated_location_first and episode:
        scene_count[episode] = scene_count.get(episode, 0) + 1
        return (
            f"{episode}-{scene_count[episode]}",
            normalize_day(generated_location_first.group("day")),
            generated_location_first.group("io"),
            clean_location(generated_location_first.group("location")),
            "",
        )
    return None


def normalize_day(day: str) -> str:
    if day in {"清晨", "上午", "下午", "傍晚"}:
        return "日"
    if day in {"深夜", "凌晨"}:
        return "夜"
    return day


def normalize_person(name: str) -> str:
    name = re.sub(r"【[^】]*】", "", name)
    name = re.sub(r"[（(][^）)]*[）)]", "", name)
    name = re.sub(r"[*＊]\s*\d+$", "", name)
    name = name.strip(" 　、，,;；:：")
    if re.fullmatch(r"[Nn\d]+", name):
        return ""
    if name in {"人物", "人物小传", "人物小传："}:
        return ""
    return name


def split_people(line: str) -> list[str]:
    line = re.sub(r"^人物\s*[:：;；]?", "", line.strip())
    line = re.sub(r"([一-龥A-Za-z]+)\s+([Nn])(?=$|[、，,\s])", r"\1\2", line)
    line = re.sub(r"([一-龥A-Za-z]+)\s+(\d+)(?=$|[、，,\s])", r"\1", line)
    parts = re.split(r"[、，,\s]+", line)
    result = []
    for part in parts:
        name = normalize_person(part)
        if name:
            result.append(name)
    return result


def augment_visible_people(scenes: list[dict]) -> None:
    candidates = ordered_unique(
        person
        for scene in scenes
        for person in scene.get("people", [])
        if not is_group_name(person)
    )
    for scene in scenes:
        people = list(scene.get("people", []))
        for role in candidates:
            if role in people:
                continue
            pattern = re.escape(role)
            for line in scene.get("body", []):
                text = line.lstrip("△").strip()
                if re.match(rf"^{pattern}VO\b", text, flags=re.IGNORECASE):
                    continue
                if re.match(rf"^{pattern}(?:[（(][^）)]*[）)])?[:：]", text) or (
                    line.startswith("△") and text.startswith(role)
                ):
                    people.append(role)
                    break
        scene["people"] = people


def parse_docx(path: Path) -> list[dict]:
    paragraphs = [p.text.strip() for p in Document(path).paragraphs if p.text.strip()]
    has_body_marker = any(p.startswith("正文") for p in paragraphs)
    in_body = not has_body_marker
    episode = None
    scene_count: dict[int, int] = {}
    scenes: list[dict] = []
    last_day = "日"

    for idx, text in enumerate(paragraphs):
        if text.startswith("正文"):
            in_body = True
            continue
        if not in_body:
            continue
        ep_match = re.match(r"^第([一二三四五六七八九十\d]+)集", text)
        if ep_match:
            episode = cn_episode_to_int(ep_match.group(1))
            if episode is not None:
                scene_count.setdefault(episode, 0)
            continue
        if not is_scene_heading(text):
            continue

        parsed = parse_heading(text, episode, scene_count, last_day)
        if not parsed:
            continue
        scene_id, day, inside_outside, location, warning = parsed
        last_day = day

        people_line = ""
        for lookahead in range(idx + 1, min(len(paragraphs), idx + 8)):
            target = paragraphs[lookahead]
            if target.startswith("人物"):
                people_line = target
                break
            if is_scene_heading(target) or re.match(r"^第[一二三四五六七八九十\d]+集", target):
                break

        scenes.append(
            {
                "source": path.name,
                "idx": idx,
                "heading": text,
                "scene": scene_id,
                "day": day,
                "inside_outside": inside_outside,
                "location": location,
                "people_raw": people_line,
                "people": split_people(people_line),
                "warning": warning,
            }
        )

    for i, scene in enumerate(scenes):
        end = scenes[i + 1]["idx"] if i + 1 < len(scenes) else len(paragraphs)
        body = []
        for text in paragraphs[scene["idx"] + 1 : end]:
            if text.startswith("人物"):
                continue
            if re.match(r"^第[一二三四五六七八九十\d]+集", text):
                continue
            body.append(text)
        scene["body"] = body
        scene["body_text"] = "\n".join(body)
    augment_visible_people(scenes)
    return scenes


def pdf_lines(path: Path) -> list[str]:
    if pdfplumber is None:
        raise RuntimeError("pdfplumber is required to parse PDF scripts")

    lines: list[str] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(
                extra_attrs=["fontname", "size", "non_stroking_color"],
                keep_blank_chars=False,
            )
            # Many script PDFs carry diagonal watermarks. These are usually
            # oversized gray words or a different watermark font; skip them
            # without deleting real Chinese characters from the script text.
            filtered = [
                word
                for word in words
                if word.get("size", 0) <= 20 and word.get("fontname") != "STSong"
            ]
            groups: list[dict] = []
            for word in filtered:
                for group in groups:
                    if abs(group["top"] - word["top"]) < 3:
                        group["words"].append(word)
                        group["top"] = (group["top"] + word["top"]) / 2
                        break
                else:
                    groups.append({"top": word["top"], "words": [word]})
            for group in sorted(groups, key=lambda item: item["top"]):
                text = "".join(
                    word["text"] for word in sorted(group["words"], key=lambda item: item["x0"])
                ).strip()
                if text:
                    lines.append(text)
    return lines


def is_pdf_heading(text: str) -> bool:
    normalized = text.replace(" ", "")
    full = r"^\d+\.\s*.+?(?:夜/日|日/夜|日|夜|清晨|上午|下午|傍晚|深夜|凌晨)(?:内外|内|外)$"
    malformed = r"^\d+\.\s*(?:（?闪回）?|（?闪现）?)?(?:一组镜头|手机屏幕|蒙太奇|转场).*$"
    return bool(re.match(full, normalized) or re.match(malformed, normalized))


def parse_pdf_heading(
    text: str,
    episode: int | str | None,
    scene_count: dict[int | str, int],
    last_day: str,
    last_inside_outside: str,
) -> tuple[str, str, str, str, str, str] | None:
    normalized = text.replace(" ", "")
    match = re.match(
        r"^(\d+)\.\s*(.+?)(夜/日|日/夜|日|夜|清晨|上午|下午|傍晚|深夜|凌晨)(内外|内|外)$",
        normalized,
    )
    warning = ""
    if match:
        source_num, location, day, inside_outside = match.groups()
        day = normalize_day(day)
    else:
        match = re.match(r"^(\d+)\.\s*(.+)$", normalized)
        if not match or not any(keyword in match.group(2) for keyword in ["一组镜头", "手机屏幕", "蒙太奇", "转场"]):
            return None
        source_num, location = match.groups()
        day = last_day or "日"
        inside_outside = "内外" if any(keyword in location for keyword in ["一组镜头", "蒙太奇", "转场"]) else (last_inside_outside or "内")
        warning = "原文未标日夜内外"

    if episode is None:
        return None
    scene_count[episode] = scene_count.get(episode, 0) + 1
    scene_id = f"{episode}-{scene_count[episode]}" if episode != "番外" else f"番外-{scene_count[episode]}"
    if scene_id.split("-")[-1] != source_num and episode != "番外":
        warning = "；".join([part for part in [warning, f"原文场号{source_num}"] if part])
    return scene_id, day, inside_outside, location.strip(), source_num, warning


def collect_pdf_people_line(lines: list[str], start: int, end: int) -> str:
    people_line = ""
    for idx in range(start + 1, min(end, start + 8)):
        text = lines[idx]
        if text.startswith("人物"):
            people_line = text
            cursor = idx + 1
            while cursor < min(end, idx + 4):
                next_text = lines[cursor].strip()
                if is_pdf_heading(next_text) or re.match(r"^第[一二三四五六七八九十\d]+集$", next_text):
                    break
                if next_text.startswith("番外篇"):
                    break
                if people_line.endswith(("、", "，", ",")) or (
                    len(next_text) <= 16
                    and "：" not in next_text
                    and ":" not in next_text
                    and not re.search(r"[。！？；;]", next_text)
                ):
                    people_line += next_text
                    cursor += 1
                    continue
                break
            break
        if is_pdf_heading(text) or re.match(r"^第[一二三四五六七八九十\d]+集$", text):
            break
    return people_line


def parse_pdf(path: Path) -> list[dict]:
    lines = pdf_lines(path)
    episode: int | str | None = None
    scene_count: dict[int | str, int] = {}
    scenes: list[dict] = []
    last_day = "日"
    last_inside_outside = "内"

    for idx, text in enumerate(lines):
        ep_match = re.match(r"^第([一二三四五六七八九十\d]+)集$", text.replace(" ", ""))
        if ep_match:
            episode = cn_episode_to_int(ep_match.group(1))
            if episode is not None:
                scene_count.setdefault(episode, 0)
            continue
        if text.replace(" ", "").startswith("番外篇"):
            episode = "番外"
            scene_count.setdefault(episode, 0)
            continue
        if not is_pdf_heading(text):
            continue
        parsed = parse_pdf_heading(text, episode, scene_count, last_day, last_inside_outside)
        if not parsed:
            continue
        scene_id, day, inside_outside, location, source_num, warning = parsed
        last_day = day
        last_inside_outside = inside_outside
        scenes.append(
            {
                "source": path.name,
                "idx": idx,
                "heading": text,
                "scene": scene_id,
                "source_num": source_num,
                "day": day,
                "inside_outside": inside_outside,
                "location": location,
                "people_raw": "",
                "people": [],
                "warning": warning,
            }
        )

    for scene_index, scene in enumerate(scenes):
        end = scenes[scene_index + 1]["idx"] if scene_index + 1 < len(scenes) else len(lines)
        people_line = collect_pdf_people_line(lines, scene["idx"], end)
        scene["people_raw"] = people_line
        scene["people"] = split_people(people_line)
        body = []
        for text in lines[scene["idx"] + 1 : end]:
            if text.startswith("人物"):
                continue
            if re.match(r"^第[一二三四五六七八九十\d]+集$", text.replace(" ", "")):
                continue
            if text.replace(" ", "").startswith("番外篇"):
                continue
            body.append(text)
        scene["body"] = body
        scene["body_text"] = "\n".join(body)
    return scenes


def parse_script(path: Path) -> list[dict]:
    if path.suffix.lower() == ".pdf":
        return parse_pdf(path)
    if path.suffix.lower() in {".docx", ".docm"}:
        return parse_docx(path)
    raise ValueError(f"Unsupported input format: {path.suffix}. Convert .doc to .docx first.")


def ordered_unique(items: list[str]) -> list[str]:
    seen = OrderedDict()
    for item in items:
        item = item.strip()
        if item:
            seen[item] = True
    return list(seen.keys())


def is_group_name(name: str) -> bool:
    return any(hint in name for hint in GROUP_HINTS)


def infer_roles(scenes: list[dict]) -> list[tuple[str, str]]:
    names = []
    for scene in scenes:
        for name in scene["people"]:
            if not is_group_name(name):
                names.append(name)
    return [(name, default_mark(name)) for name in ordered_unique(names)]


def default_mark(role: str) -> str:
    if role.endswith("老爷子"):
        return "爷"
    if role.endswith("助理"):
        return role[0]
    if role.startswith("小") and len(role) >= 2:
        return role[-1]
    cleaned = re.sub(r"[（）()青年]", "", role)
    return cleaned[-1] if cleaned else role[-1]


def short_summary(scene: dict) -> str:
    fragments = []
    for line in scene["body"]:
        line = re.sub(r"^△", "", line).strip()
        line = re.sub(r"【[^】]*】", "", line)
        line = re.sub(r"[（(].*?[）)]", "", line)
        line = re.sub(r"^\w+[:：；;]", "", line).strip()
        line = re.sub(r"[，。！？、\s]+", "", line)
        if line:
            fragments.append(line)
        candidate = "".join(fragments)
        if len(candidate) >= 12:
            return candidate[:18]
    if fragments:
        return "".join(fragments)[:18]
    return "本场拍摄内容需要人工补写"


def collect_groups(scene: dict) -> str:
    groups = []
    text = scene["people_raw"] + "\n" + scene["body_text"]
    for person in scene["people"]:
        if is_group_name(person):
            groups.append(re.sub(r"(若干|[Nn]|\d+)$", "", person))
    if re.search(r"众人|所有人|全场|会议室里的人|各位", text):
        groups.append("会议人员" if "会议" in scene["location"] else "众人")
    if re.search(r"应聘者|面试者", text):
        groups.append("应聘者")
    if "保镖" in text:
        groups.append("保镖")
    if re.search(r"员工|同事|职员", text):
        groups.append("公司员工")
    return "、".join(ordered_unique(groups))


def collect_props(scene: dict) -> str:
    text = scene["body_text"]
    props = []
    for keyword, label in PROP_KEYWORDS:
        if keyword == "刀" and not re.search(r"持刀|抽出.*刀|举刀|刀刃|刀身|刀尖|短刀|菜刀|水果刀|美工刀|手中.*刀|刀.*砍|刀.*刺", text):
            continue
        if keyword in text:
            props.append(label)
    if "头发" in text and re.search(r"拔了几根|扯了几根|亲子鉴定", text):
        props.append("头发样本")
    props = ordered_unique(props)
    if "水果刀" in props and "刀具" in props:
        props.remove("刀具")
    if "美工刀" in props and "刀具" in props:
        props.remove("刀具")
    if "子弹/枪械效果" in props and "枪械效果" in props:
        props.remove("枪械效果")
    if "手机" in props and "手机/电话" in props:
        props.remove("手机/电话")
    return "、".join(props[:8])


def collect_costume(scene: dict) -> str:
    text = scene["people_raw"] + "\n" + scene["body_text"]
    hints = []
    if "青年" in text:
        hints.append("青年妆造")
    if re.search(r"怀孕|孕吐|反胃|摸了摸肚子", text):
        hints.append("孕期状态")
    if re.search(r"湿透|滴水|泼冷水|浑身冰冷|嘴唇发紫", text):
        hints.append("湿发湿衣/虚弱妆")
    if re.search(r"包扎|受伤|划伤|伤口|带血|乌青|红肿|昏迷|苍白|惨白|虚弱|憔悴|病床|受凉|发烧", text):
        hints.append("病弱/伤口妆")
    if "衣衫不整" in text:
        hints.append("衣衫不整")
    if re.search(r"西装|高定", text):
        hints.append("西装")
    if "职业套装" in text:
        hints.append("职业套装")
    if "连衣裙" in text:
        hints.append("连衣裙")
    if "洗得发白" in text or "旧外套" in text:
        hints.append("朴素旧衣")
    if "保洁" in text or "清洁工" in text:
        hints.append("保洁/朴素造型")
    if "衣衫褴褛" in text:
        hints.append("狼狈造型")
    return "、".join(ordered_unique(hints))


def collect_notes(scene: dict) -> str:
    text = scene["body_text"]
    notes = []
    if "字幕" in text:
        notes.append("含字幕")
    if re.search(r"VO|vo|画外音", text):
        notes.append("含VO")
    if re.search(r"闪回|回忆", text):
        notes.append("闪回/回忆")
    if re.search(r"接吻|亲吻|强吻|乱亲|床上|压在|抱住.*亲|衣衫不整|睡裙|钻进被窝", text):
        notes.append("亲密戏注意尺度")
    if re.search(r"扭打|持刀|水果刀|美工刀|踹开|砸|扑倒|推倒|滚远|撞在地上|拽|拉扯|摔", text):
        notes.append("动作戏注意安全")
    if "接" in scene["location"]:
        notes.append(re.sub(r".*?(接[^）)]*).*", r"\1", scene["location"]))
    if "一同拍摄" in scene["location"]:
        notes.append("一同拍摄")
    if scene.get("warning"):
        notes.append(scene["warning"])
    return "；".join(ordered_unique(notes))


def vertical(text: str) -> str:
    return "\n".join(text)


def load_config(path: Path | None) -> dict:
    if not path:
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def build_workbook(source: Path, scenes: list[dict], output: Path, config: dict) -> None:
    role_config = config.get("roles")
    roles = [(item["name"], item.get("mark") or default_mark(item["name"])) for item in role_config] if role_config else infer_roles(scenes)
    summaries = config.get("summaries", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "顺场表"
    headers = ["拍摄顺序", "实际拍摄场地", "场次", "剧本中场景", "拍摄内容"]
    headers += ["日\n/\n夜", "内\n/\n外", "页\n数"]
    headers += [vertical(role) for role, _ in roles]
    headers += ["群演", "梳化服提示", "道具提示", "备注/特殊道具/时间"]
    ws.append(headers)

    role_names = {role for role, _ in roles}
    for order, scene in enumerate(scenes, start=1):
        people = set(scene["people"])
        row = [
            order,
            "",
            scene["scene"],
            scene["location"],
            summaries.get(scene["scene"], short_summary(scene)),
            scene["day"],
            scene["inside_outside"],
            "",
        ]
        for role, mark in roles:
            row.append(mark if role in people and role in role_names else "")
        row += [
            collect_groups(scene),
            collect_costume(scene),
            collect_props(scene),
            collect_notes(scene),
        ]
        ws.append(row)

    style_sheet(ws, len(headers), len(scenes), len(roles))
    wb.save(output)


def style_sheet(ws, max_col: int, scene_count: int, role_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    header_font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=10)
    body_font = Font(name="Microsoft YaHei", color="222222", size=10)
    alt_fill = PatternFill("solid", fgColor="F4F8FA")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="B8C7CE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{scene_count + 1}"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 82

    for row in range(2, scene_count + 2):
        fill = alt_fill if row % 2 == 0 else white_fill
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in [4, 5, max_col - 3, max_col - 2, max_col - 1, max_col]:
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 36

    for col, width in {1: 8, 2: 14, 3: 9, 4: 23, 5: 20}.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    for col in range(6, 9):
        ws.column_dimensions[get_column_letter(col)].width = 4.2
    role_start = 9
    role_end = role_start + role_count - 1
    for col in range(role_start, role_end + 1):
        ws.column_dimensions[get_column_letter(col)].width = 4.2
    ws.column_dimensions[get_column_letter(role_end + 1)].width = 15
    ws.column_dimensions[get_column_letter(role_end + 2)].width = 20
    ws.column_dimensions[get_column_letter(role_end + 3)].width = 22
    ws.column_dimensions[get_column_letter(role_end + 4)].width = 26


def title_from_path(path: Path) -> str:
    name = path.stem
    match = re.search(r"《([^》]+)》", name)
    return match.group(1) if match else name


def dump_json(path: Path, all_scenes: dict[str, list[dict]]) -> None:
    serializable = {}
    for key, scenes in all_scenes.items():
        serializable[key] = [
            {k: v for k, v in scene.items() if k not in {"body_text"}}
            for scene in scenes
        ]
    path.write_text(json.dumps(serializable, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build short-drama 顺场表 Excel files from DOCX/PDF scripts.")
    parser.add_argument("scripts", nargs="+", type=Path)
    parser.add_argument("--out-dir", type=Path, default=Path.cwd())
    parser.add_argument("--config", type=Path)
    parser.add_argument("--dump-json", type=Path)
    args = parser.parse_args()

    config = load_config(args.config)
    args.out_dir.mkdir(parents=True, exist_ok=True)
    all_scenes = {}

    for source in args.scripts:
        scenes = parse_script(source)
        all_scenes[source.name] = scenes
        output = args.out_dir / f"顺场表-{title_from_path(source)}.xlsx"
        per_file_config = config.get(source.name, config)
        build_workbook(source, scenes, output, per_file_config)
        warnings = [f"{s['scene']}:{s['warning']}" for s in scenes if s.get("warning")]
        print(f"{source.name}: {len(scenes)} scenes -> {output}")
        if warnings:
            print("  warnings:", ", ".join(warnings))

    if args.dump_json:
        dump_json(args.dump_json, all_scenes)
        print(f"scene dump -> {args.dump_json}")


if __name__ == "__main__":
    main()
